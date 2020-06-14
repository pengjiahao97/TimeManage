from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet


class MyStyle:
    # 字体
    font = Font(
        name="Consolas",
        size=11
    )
    # 对齐方式
    alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    # 边框
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'),
    )

    @staticmethod
    def solid_fill(hex_color_code):
        """纯色背景填充"""
        return PatternFill(fill_type="solid", start_color=hex_color_code)

    @staticmethod
    def adjust_height_and_width(worksheet: Worksheet, height, width):
        """调整sheet页单元格的高和宽"""
        for row in range(worksheet.max_row):
            worksheet.row_dimensions[row + 1].height = height
        for col in range(worksheet.max_column):
            col_letter = get_column_letter(col + 1)
            worksheet.column_dimensions[col_letter].width = width
