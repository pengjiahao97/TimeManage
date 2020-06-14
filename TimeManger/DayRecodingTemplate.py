import os
from openpyxl import Workbook, load_workbook
from TimeManger import MyStyle
from datetime import datetime


def generate_day_recoding_excel(excel_path):
    if not os.path.isabs(excel_path):
        excel_path = os.path.abspath(excel_path)

    now = datetime.now().strftime("%Y-%m-%d")
    sheet_name = f"day_{now}"
    first_row_cn = ("时间", "预期结果", "实际结果", "是否达到预期")

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.create_sheet(sheet_name, 0)
        print("asdasd")
    else:
        print("asdasdasd")
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    ws.cell(10, 4).value = ""  # 初始化10行单元格

    for i in range(4):
        ws.cell(1, i + 1).value = first_row_cn[i]

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.font = MyStyle.font
            cell.alignment = MyStyle.alignment
            cell.border = MyStyle.border

    MyStyle.adjust_height_and_width(ws, 15, 13.5)
    wb.save(excel_path)
