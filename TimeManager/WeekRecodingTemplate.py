import os
from openpyxl import Workbook, load_workbook
from TimeManager import MyStyle


def generate_week_recoding_excel(excel_path):
    if not os.path.isabs(excel_path):
        excel_path = os.path.abspath(excel_path)

    sheet_name = "week"
    first_col_field = "Time"
    workday_cn = {1: "星期一", 2: "星期二", 3: "星期三", 4: "星期四", 5: "星期五"}

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.create_sheet(sheet_name, 0)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    ws.cell(1, 1).value = first_col_field

    green_cell, red_cell, purple_cell = ws.cell(2, 8, "集中精力"), ws.cell(3, 8, "浪费时间"), ws.cell(4, 8, "放松休息")
    green_cell.fill, green_cell.border = MyStyle.solid_fill("92D050"), MyStyle.border  # green: focus on working
    red_cell.fill, red_cell.border = MyStyle.solid_fill("FF0000"), MyStyle.border  # red: waste time for no meaning
    purple_cell.fill, purple_cell.border = MyStyle.solid_fill("B1A0C7"), MyStyle.border  # purple: rest

    for i in range(2, 16):
        ws.cell(i, 1).value = f"{i + 6}点"
    for i in range(2, 7):
        ws.cell(1, i).value = workday_cn.get(i - 1)  # 工作日

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.font = MyStyle.font
            cell.alignment = MyStyle.alignment
            if cell.column != 7:
                cell.border = MyStyle.border

    MyStyle.adjust_height_and_width(ws, 15, 8)
    wb.save(excel_path)
